# 依赖库最低版本要求: pandas>=2.0, openpyxl>=3.1
# 数据来源: 本文件使用 pandas 构造的模拟数据集，无需外部数据文件

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter

output_dir = Path(__file__).parent / "output"
output_dir.mkdir(exist_ok=True)

print("=" * 60)
print("2. Excel多Sheet与格式化输出")
print("=" * 60)

np.random.seed(42)
df = pd.DataFrame({
    "姓名": [f"员工{i:02d}" for i in range(1, 21)],
    "部门": np.random.choice(["技术部", "市场部", "财务部", "人事部"], 20),
    "月薪": np.random.randint(8000, 25000, 20),
    "绩效评分": np.round(np.random.uniform(60, 100, 20), 1),
    "入职日期": pd.date_range("2022-01-01", periods=20, freq="30D"),
})

# --------------------------------------------------
# 一、使用 openpyxl 创建多 Sheet 工作簿
# --------------------------------------------------

wb = Workbook()

ws1 = wb.active
ws1.title = "员工明细"

headers = ["姓名", "部门", "月薪", "绩效评分", "入职日期"]
for col_idx, header in enumerate(headers, 1):
    ws1.cell(row=1, column=col_idx, value=header)

for row_idx, row in df.iterrows():
    for col_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=row_idx + 2, column=col_idx, value=value)
        if col_idx == 5:
            cell.number_format = "YYYY-MM-DD"

print("[Sheet1] 员工明细 - 数据已写入")

ws2 = wb.create_sheet(title="部门汇总")
dept_summary = df.groupby("部门").agg(
    人数=("姓名", "count"),
    平均月薪=("月薪", "mean"),
    平均绩效=("绩效评分", "mean"),
).round(2)

summary_headers = ["部门", "人数", "平均月薪", "平均绩效"]
for col_idx, header in enumerate(summary_headers, 1):
    ws2.cell(row=1, column=col_idx, value=header)

for row_idx, (dept, row) in enumerate(dept_summary.iterrows(), 2):
    ws2.cell(row=row_idx, column=1, value=dept)
    for col_idx, value in enumerate(row, 2):
        ws2.cell(row=row_idx, column=col_idx, value=value)

print("[Sheet2] 部门汇总 - 数据已写入")

ws3 = wb.create_sheet(title="薪资分布")
bins = [8000, 12000, 16000, 20000, 25000]
labels = ["8K-12K", "12K-16K", "16K-20K", "20K-25K"]
df["薪资区间"] = pd.cut(df["月薪"], bins=bins, labels=labels, right=True)
salary_dist = df["薪资区间"].value_counts().sort_index()

ws3.cell(row=1, column=1, value="薪资区间")
ws3.cell(row=1, column=2, value="人数")
for row_idx, (interval, count) in enumerate(salary_dist.items(), 2):
    ws3.cell(row=row_idx, column=1, value=str(interval))
    ws3.cell(row=row_idx, column=2, value=count)

print("[Sheet3] 薪资分布 - 数据已写入")

# --------------------------------------------------
# 二、单元格样式设置
# --------------------------------------------------

header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

for ws in [ws1, ws2, ws3]:
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

data_font = Font(name="微软雅黑", size=10)
data_alignment = Alignment(horizontal="center", vertical="center")

for ws in [ws1, ws2, ws3]:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

print("\n[样式] 表头和数据单元格样式已应用")

# --------------------------------------------------
# 三、条件格式
# --------------------------------------------------

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_font = Font(color="006100")
red_font = Font(color="9C0006")

ws1.conditional_formatting.add(
    "C2:C21",
    CellIsRule(operator="greaterThan", formula=["20000"], fill=green_fill, font=green_font),
)
ws1.conditional_formatting.add(
    "C2:C21",
    CellIsRule(operator="lessThan", formula=["12000"], fill=red_fill, font=red_font),
)
print("[条件格式] 月薪: >20000绿色, <12000红色")

ws1.conditional_formatting.add(
    "D2:D21",
    DataBarRule(start_type="min", end_type="max", color="5B9BD5"),
)
print("[条件格式] 绩效评分: 数据条")

# --------------------------------------------------
# 四、数字格式
# --------------------------------------------------

for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=3, max_col=3):
    for cell in row:
        cell.number_format = '#,##0"元"'

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=3, max_col=3):
    for cell in row:
        cell.number_format = '#,##0.00"元"'

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=4, max_col=4):
    for cell in row:
        cell.number_format = "0.00"

print("[数字格式] 薪资: 千分位+元, 绩效: 两位小数")

# --------------------------------------------------
# 五、列宽自适应
# --------------------------------------------------

for ws in [ws1, ws2, ws3]:
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        adjusted_width = max(max_length + 4, 12)
        ws.column_dimensions[col_letter].width = adjusted_width

print("[列宽] 已自适应调整")

ws1.auto_filter.ref = f"A1:E{ws1.max_row}"
ws1.freeze_panes = "A2"

formatted_path = output_dir / "格式化员工报表.xlsx"
wb.save(formatted_path)
print(f"\n[保存] 格式化报表已保存: {formatted_path}")

# --------------------------------------------------
# 六、使用 pandas + openpyxl 快捷方式
# --------------------------------------------------

excel_path = output_dir / "pandas格式化报表.xlsx"
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="员工数据", index=False)
    dept_summary.to_excel(writer, sheet_name="部门汇总")

    workbook = writer.book
    ws = workbook["员工数据"]

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

print(f"[pandas+openpyxl] 快捷格式化报表已保存: {excel_path}")

print("\n" + "=" * 60)
print("Excel多Sheet与格式化输出 - 完成")
print("=" * 60)
