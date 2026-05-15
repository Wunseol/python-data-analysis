# 依赖库最低版本要求: pandas>=2.0, openpyxl>=3.1, fpdf2>=2.7, matplotlib>=3.7
# 数据来源: 本文件使用 pandas 构造的模拟数据集，无需外部数据文件
# 说明: 本案例演示完整的自动化周报生成流程，涵盖数据生成→分析→图表→Excel→PDF→Markdown

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei"]
plt.rcParams["axes.unicode_minus"] = False

output_dir = Path(__file__).parent / "output" / "weekly_report"
output_dir.mkdir(parents=True, exist_ok=True)

print("=" * 60)
print("8. 综合案例 - 自动化周报生成")
print("=" * 60)

# ==================================================
# 第一步: 生成模拟数据
# ==================================================

print("\n[步骤1] 生成模拟数据...")

np.random.seed(42)
departments = ["技术部", "市场部", "财务部", "人事部"]
current_date = datetime.now()

df_employees = pd.DataFrame({
    "姓名": [f"员工{i:02d}" for i in range(1, 31)],
    "部门": np.random.choice(departments, 30),
    "月薪": np.random.randint(8000, 30000, 30),
    "绩效评分": np.round(np.random.uniform(55, 100, 30), 1),
    "入职日期": pd.date_range("2021-01-01", periods=30, freq="20D"),
})

dates = pd.date_range(end=current_date, periods=7, freq="D")
df_daily = pd.DataFrame({
    "日期": np.repeat(dates, 5),
    "部门": np.tile(departments, len(dates))[: len(dates) * 5] if len(dates) * 5 <= 35 else np.random.choice(departments, len(dates) * 5),
    "完成任务数": np.random.randint(3, 20, len(dates) * 5 if len(dates) * 5 <= 35 else 35),
    "工时": np.round(np.random.uniform(6, 10, len(dates) * 5 if len(dates) * 5 <= 35 else 35), 1),
})

num_daily = len(dates) * 5
df_daily = pd.DataFrame({
    "日期": np.repeat(dates, 5),
    "部门": np.random.choice(departments, num_daily),
    "完成任务数": np.random.randint(3, 20, num_daily),
    "工时": np.round(np.random.uniform(6, 10, num_daily), 1),
})

print(f"  员工数据: {df_employees.shape}")
print(f"  每日数据: {df_daily.shape}")

# ==================================================
# 第二步: 数据分析
# ==================================================

print("\n[步骤2] 数据分析...")

dept_summary = df_employees.groupby("部门").agg(
    人数=("姓名", "count"),
    平均月薪=("月薪", "mean"),
    平均绩效=("绩效评分", "mean"),
    最高月薪=("月薪", "max"),
    最低月薪=("月薪", "min"),
).round(2)

daily_dept = df_daily.groupby("部门").agg(
    总完成任务=("完成任务数", "sum"),
    总工时=("工时", "sum"),
    平均日工时=("工时", "mean"),
).round(2)

daily_trend = df_daily.groupby("日期").agg(
    总任务数=("完成任务数", "sum"),
    总工时=("工时", "sum"),
).round(1)

top_performers = df_employees.nlargest(5, "绩效评分")[["姓名", "部门", "绩效评分", "月薪"]]
low_performers = df_employees.nsmallest(3, "绩效评分")[["姓名", "部门", "绩效评分", "月薪"]]

print(f"  部门汇总: {dept_summary.shape}")
print(f"  每日趋势: {daily_trend.shape}")

# ==================================================
# 第三步: 生成图表
# ==================================================

print("\n[步骤3] 生成图表...")

fig, axes = plt.subplots(2, 2, figsize=(14, 10))

ax = axes[0, 0]
dept_avg_salary = dept_summary["平均月薪"].sort_values()
bars = ax.barh(dept_avg_salary.index, dept_avg_salary.values, color="#4472C4")
for bar, val in zip(bars, dept_avg_salary.values):
    ax.text(val + 200, bar.get_y() + bar.get_height() / 2, f"{val:.0f}", va="center")
ax.set_title("各部门平均月薪")
ax.set_xlabel("月薪 (元)")

ax = axes[0, 1]
dept_avg_perf = dept_summary["平均绩效"].sort_values()
colors = ["#FF6B6B" if v < 75 else "#4ECDC4" for v in dept_avg_perf.values]
ax.bar(dept_avg_perf.index, dept_avg_perf.values, color=colors)
ax.set_title("各部门平均绩效评分")
ax.set_ylabel("绩效评分")
ax.axhline(y=df_employees["绩效评分"].mean(), color="red", linestyle="--", alpha=0.5, label="整体均值")
ax.legend()

ax = axes[1, 0]
ax.plot(daily_trend.index, daily_trend["总任务数"], "o-", color="#4472C4", linewidth=2, markersize=6)
ax.fill_between(daily_trend.index, daily_trend["总任务数"], alpha=0.15, color="#4472C4")
ax.set_title("每日完成任务数趋势")
ax.set_ylabel("任务数")
ax.tick_params(axis="x", rotation=45)

ax = axes[1, 1]
dept_tasks = daily_dept["总完成任务"]
ax.pie(dept_tasks, labels=dept_tasks.index, autopct="%1.1f%%", startangle=90,
       colors=["#4472C4", "#ED7D31", "#A5A5A5", "#FFC000"])
ax.set_title("各部门任务占比")

plt.suptitle(f"周报数据分析 ({current_date.strftime('%Y-%m-%d')})", fontsize=14, fontweight="bold")
plt.tight_layout()
chart_path = output_dir / "周报图表.png"
fig.savefig(chart_path, dpi=150, bbox_inches="tight")
plt.close(fig)
print(f"  图表已保存: {chart_path}")

# ==================================================
# 第四步: 生成 Excel 汇总
# ==================================================

print("\n[步骤4] 生成Excel汇总...")

wb = Workbook()

header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
data_font = Font(name="微软雅黑", size=10)
data_align = Alignment(horizontal="center", vertical="center")

def style_header(ws, num_cols):
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_data(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

def auto_width(ws):
    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(cell.value or "")) for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 4, 12)

ws1 = wb.active
ws1.title = "员工明细"
for col_idx, header in enumerate(df_employees.columns, 1):
    ws1.cell(row=1, column=col_idx, value=header)
for row_idx, row in enumerate(df_employees.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws1.cell(row=row_idx, column=col_idx, value=value)
style_header(ws1, len(df_employees.columns))
style_data(ws1)
auto_width(ws1)

ws2 = wb.create_sheet("部门汇总")
for col_idx, header in enumerate(dept_summary.reset_index().columns, 1):
    ws2.cell(row=1, column=col_idx, value=header)
for row_idx, row in enumerate(dept_summary.reset_index().itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)
style_header(ws2, len(dept_summary.columns) + 1)
style_data(ws2)
auto_width(ws2)

ws3 = wb.create_sheet("每日趋势")
for col_idx, header in enumerate(daily_trend.reset_index().columns, 1):
    ws3.cell(row=1, column=col_idx, value=header)
for row_idx, row in enumerate(daily_trend.reset_index().itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws3.cell(row=row_idx, column=col_idx, value=value)
style_header(ws3, len(daily_trend.columns) + 1)
style_data(ws3)
auto_width(ws3)

excel_path = output_dir / f"周报Excel_{current_date.strftime('%Y%m%d')}.xlsx"
wb.save(excel_path)
print(f"  Excel已保存: {excel_path}")

# ==================================================
# 第五步: 生成 PDF 报告
# ==================================================

print("\n[步骤5] 生成PDF报告...")

class WeeklyPDF(FPDF):
    def __init__(self):
        super().__init__()
        self.cn_font = "helvetica"

    def header(self):
        self.set_font(self.cn_font, "B", 9)
        self.set_text_color(100, 100, 100)
        self.cell(0, 6, f"自动化周报 | {current_date.strftime('%Y-%m-%d')}", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(47, 84, 150)
        self.set_line_width(0.3)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(3)

    def footer(self):
        self.set_y(-12)
        self.set_font(self.cn_font, "", 7)
        self.set_text_color(128)
        self.cell(0, 8, f"Page {self.page_no()}/{{nb}}", align="C")

    def section_title(self, title):
        self.set_font(self.cn_font, "B", 14)
        self.set_text_color(47, 84, 150)
        self.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
        self.ln(2)

    def body_text(self, text):
        self.set_font(self.cn_font, "", 10)
        self.set_text_color(0)
        self.multi_cell(0, 7, text)
        self.ln(2)

    def add_table(self, headers, data, col_widths):
        self.set_font(self.cn_font, "B", 9)
        self.set_fill_color(47, 84, 150)
        self.set_text_color(255, 255, 255)
        for i, header in enumerate(headers):
            self.cell(col_widths[i], 8, header, border=1, fill=True, align="C")
        self.ln()

        self.set_font(self.cn_font, "", 9)
        self.set_text_color(0)
        fill = False
        for row in data:
            if fill:
                self.set_fill_color(235, 241, 250)
            else:
                self.set_fill_color(255, 255, 255)
            for i, value in enumerate(row):
                self.cell(col_widths[i], 7, str(value), border=1, fill=True, align="C")
            self.ln()
            fill = not fill

pdf = WeeklyPDF()
pdf.alias_nb_pages()
pdf.set_auto_page_break(auto=True, margin=15)

pdf.add_page()
pdf.set_font(pdf.cn_font, "B", 22)
pdf.ln(30)
pdf.cell(0, 12, "Weekly Data Report", align="C", new_x="LMARGIN", new_y="NEXT")
pdf.set_font(pdf.cn_font, "", 12)
pdf.set_text_color(100)
pdf.ln(5)
pdf.cell(0, 8, f"Report Date: {current_date.strftime('%Y-%m-%d')}", align="C", new_x="LMARGIN", new_y="NEXT")
pdf.cell(0, 8, f"Data Period: Last 7 Days", align="C", new_x="LMARGIN", new_y="NEXT")
pdf.cell(0, 8, f"Total Employees: {len(df_employees)}", align="C", new_x="LMARGIN", new_y="NEXT")

pdf.add_page()
pdf.section_title("1. Executive Summary")
pdf.body_text(
    f"This report covers {len(df_employees)} employees across {df_employees['部门'].nunique()} departments. "
    f"Average salary: {df_employees['月薪'].mean():,.0f} CNY. "
    f"Average performance: {df_employees['绩效评分'].mean():.1f}. "
    f"Total tasks completed this week: {df_daily['完成任务数'].sum()}."
)

pdf.section_title("2. Department Summary")
headers = ["Department", "Headcount", "Avg Salary", "Avg Perf"]
data = []
for dept, row in dept_summary.iterrows():
    data.append([dept, int(row["人数"]), f"{row['平均月薪']:,.0f}", f"{row['平均绩效']:.1f}"])
pdf.add_table(headers, data, [40, 30, 45, 45])

pdf.ln(5)
pdf.section_title("3. Top Performers")
headers = ["Name", "Department", "Performance", "Salary"]
data = []
for _, row in top_performers.iterrows():
    data.append([row["姓名"], row["部门"], f"{row['绩效评分']:.1f}", f"{row['月薪']:,}"])
pdf.add_table(headers, data, [35, 35, 40, 40])

pdf.add_page()
pdf.section_title("4. Visualizations")
pdf.image(str(chart_path), x=10, w=190)

pdf_path = output_dir / f"周报PDF_{current_date.strftime('%Y%m%d')}.pdf"
pdf.output(str(pdf_path))
print(f"  PDF已保存: {pdf_path}")

# ==================================================
# 第六步: 生成 Markdown 摘要
# ==================================================

print("\n[步骤6] 生成Markdown摘要...")

report_date = current_date.strftime("%Y年%m月%d日")
week_start = (current_date - timedelta(days=6)).strftime("%m月%d日")
week_end = current_date.strftime("%m月%d日")

md_content = f"""# 自动化周报摘要

> 生成时间: {report_date}
> 数据周期: {week_start} - {week_end}

## 关键指标

| 指标 | 数值 |
|------|------|
| 员工总数 | {len(df_employees)} 人 |
| 平均月薪 | {df_employees['月薪'].mean():,.0f} 元 |
| 平均绩效 | {df_employees['绩效评分'].mean():.1f} 分 |
| 本周完成任务 | {df_daily['完成任务数'].sum()} 个 |
| 本周总工时 | {df_daily['工时'].sum():.1f} 小时 |

## 部门汇总

| 部门 | 人数 | 平均月薪 | 平均绩效 | 本周任务数 |
|------|------|---------|---------|-----------|
"""

for dept, row in dept_summary.iterrows():
    tasks = daily_dept.loc[dept, "总完成任务"] if dept in daily_dept.index else 0
    md_content += f"| {dept} | {int(row['人数'])} | {row['平均月薪']:,.0f} | {row['平均绩效']:.1f} | {tasks} |\n"

md_content += f"""
## 绩效TOP5

| 排名 | 姓名 | 部门 | 绩效评分 | 月薪 |
|------|------|------|---------|------|
"""

for rank, (_, row) in enumerate(top_performers.iterrows(), 1):
    md_content += f"| {rank} | {row['姓名']} | {row['部门']} | {row['绩效评分']:.1f} | {row['月薪']:,} |\n"

md_content += f"""
## 需关注员工

| 姓名 | 部门 | 绩效评分 | 月薪 |
|------|------|---------|------|
"""

for _, row in low_performers.iterrows():
    md_content += f"| {row['姓名']} | {row['部门']} | {row['绩效评分']:.1f} | {row['月薪']:,} |\n"

md_content += f"""
## 可视化

![周报图表]({chart_path.name})

## 结论与建议

1. 本周共完成 **{df_daily['完成任务数'].sum()}** 项任务，总工时 **{df_daily['工时'].sum():.1f}** 小时
2. 绩效评分最高部门: **{dept_summary['平均绩效'].idxmax()}** ({dept_summary['平均绩效'].max():.1f}分)
3. 薪资最高部门: **{dept_summary['平均月薪'].idxmax()}** (平均{dept_summary['平均月薪'].max():,.0f}元)
4. 共有 **{len(low_performers)}** 名员工绩效偏低，建议关注

---

*本报告由 Python 自动化周报系统生成*
*生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*
"""

md_path = output_dir / f"周报摘要_{current_date.strftime('%Y%m%d')}.md"
md_path.write_text(md_content, encoding="utf-8")
print(f"  Markdown已保存: {md_path}")

# ==================================================
# 第七步: 输出汇总
# ==================================================

print("\n" + "=" * 60)
print("自动化周报生成 - 完成")
print("=" * 60)

print(f"\n生成文件清单:")
for f in sorted(output_dir.rglob("*")):
    if f.is_file():
        size_kb = f.stat().st_size / 1024
        print(f"  {f.name} ({size_kb:.1f} KB)")

print(f"\n输出目录: {output_dir}")
print("\n完整自动化流程:")
print("  1. 生成模拟数据 (员工+每日)")
print("  2. 数据分析 (部门汇总/每日趋势/TOP排名)")
print("  3. 生成可视化图表 (4子图)")
print("  4. 生成Excel汇总 (3个Sheet)")
print("  5. 生成PDF报告 (封面+摘要+表格+图表)")
print("  6. 生成Markdown摘要 (完整指标+表格+结论)")
